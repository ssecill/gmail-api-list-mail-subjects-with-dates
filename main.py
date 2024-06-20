import os
import json
from datetime import datetime, timedelta
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pytz

# Yetkilendirme kapsamları
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
bugun = datetime.now()

def authenticate_gmail():
    creds = None
    token_path = 'token.json'

    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = Credentials.from_authorized_user_info(json.load(token), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
            with open(token_path, 'w') as token:
                token.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)

def list_messages(service, user_id, query=''):
    try:
        response = service.users().messages().list(userId=user_id, q=query).execute()
        messages = []
        if 'messages' in response:
            messages.extend(response['messages'])
        while 'nextPageToken' in response:
            page_token = response['nextPageToken']
            response = service.users().messages().list(userId=user_id, q=query, pageToken=page_token).execute()
            messages.extend(response['messages'])
        return messages
    except Exception as error:
        print(f'An error occurred: {error}')
        return []

def get_message_details(service, user_id, msg_id):
    try:
        message = service.users().messages().get(userId=user_id, id=msg_id, format='metadata',
                                                 metadataHeaders=['subject', 'date']).execute()
        headers = message['payload']['headers']
        subject = None
        date = None
        for header in headers:
            if header['name'] == 'Subject':
                subject = header['value']
            if header['name'] == 'Date':
                date = header['value']
        return subject, date
    except Exception as error:
        print(f'An error occurred: {error}')
        return None, None

def adjust_column_width_and_wrap_text(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.alignment = Alignment(wrap_text=True)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def main():
    service = authenticate_gmail()

    # Tarih aralığı belirleme
    start_date = bugun.strftime('%Y/%m/%d') #'2024/06/20'  # Başlangıç tarihi
    end_date = (bugun + timedelta(days=1)).strftime('%Y/%m/%d') # Bitiş tarihi

    # Gmail arama sorgusu
    query = f'label:aws-alarm after:{start_date} before:{end_date}'

    # Mesajları listeleme
    messages = list_messages(service, 'me', query=query)

    if not messages:
        print('No messages found.')
    else:
        email_data = []
        for message in messages:
            msg_id = message['id']
            subject, date = get_message_details(service, 'me', msg_id)
            if subject and date and "error_alarm" in subject:
                date_parsed = datetime.strptime(date, '%a, %d %b %Y %H:%M:%S %z').replace(tzinfo=None)
                email_data.append([date_parsed, subject])

        if email_data:
            df = pd.DataFrame(email_data, columns=['DateTime', 'Subject'])
            
            # Set date format explicitly
            df['DateTime'] = df['DateTime'].dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Yerel saat dilimi kolonu ekle
            local_tz = pytz.timezone('Europe/Istanbul')  # Yerel saat dilimini belirtin
            df['LocalDateTime'] = pd.to_datetime(df['DateTime']).apply(lambda x: pytz.utc.localize(x).astimezone(local_tz).strftime('%Y-%m-%d %H:%M:%S'))
            
            print(df)
            print(f'Total Emails: {len(email_data)}')
            
            # Save the dataframe to an Excel file
            file_path = 'email_data.xlsx'
            df.to_excel(file_path, index=False)

            # Adjust column width and wrap text
            wb = load_workbook(file_path)
            ws = wb.active
            adjust_column_width_and_wrap_text(ws)
            wb.save(file_path)
            print('Data saved to email_data.xlsx')
        else:
            print('No valid messages found.')

if __name__ == '__main__':
    main()
